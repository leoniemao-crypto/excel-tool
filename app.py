import os
import sys
import subprocess
import importlib

# 🚀 終極防錯：在網頁資料夾內建立專屬工具箱，強制繞過雲端系統安裝 Bug
local_libs = os.path.join(os.getcwd(), "local_packages")
if local_libs not in sys.path:
    sys.path.insert(0, local_libs)

for package, import_name in [("openpyxl", "openpyxl"), ("Pillow", "PIL")]:
    try:
        importlib.import_module(import_name)
    except ImportError:
        subprocess.check_call([sys.executable, "-m", "pip", "install", "--target", local_libs, package])
        importlib.invalidate_caches()

# --- LINE Pay 票券雙效終極大平台主程式 ---
import streamlit as st
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from PIL import Image
from datetime import datetime, timedelta
import io
import zipfile  # 內建庫，用於多張商品圖打包下載

st.set_page_config(page_title="LINE Pay 新規大平台", layout="wide")

if "clear_id" not in st.session_state:
    st.session_state["clear_id"] = 0

cid = st.session_state["clear_id"]

# 建立兩大服務分頁
tab1, tab2 = st.tabs(["📝 1. 票券純文字 Excel 生成", "🖼️ 2. 官方圖片尺寸調整工具"])

# ==========================================
# 📝 分頁一：純文字 Excel 生成邏輯
# ==========================================
with tab1:
    st.title("🚀 LINE Pay 票券大量商品上架標準 Excel 生成器")
    st.error("💡 官方規格提醒：不同規格的商品需要分開上架（例如：大、中、小杯，或是冷、熱飲，皆須拆分為獨立的商品品項填寫）。")
    
    st.write("---")
    st.subheader("🏢 實際商品(服務)提供者基本資料填寫")
    c1, c2, c3 = st.columns(3)
    with c1:
        brand_name = st.text_input("品牌名稱 (兼商店名稱)", placeholder="例如：笨道策展咖啡", key=f"brand_name_{cid}")
        store_tax_id = st.text_input("商店統編", value="24941093", key=f"store_tax_id_{cid}")
    with c2:
        store_address = st.text_input("商店地址", value="台北市南港區", key=f"store_address_{cid}")
        store_phone1 = st.text_input("商店電話 1", value="0912345678", key=f"store_phone1_{cid}")
    with c3:
        store_phone2 = st.text_input("商店電話 2", value="02-6631-5166", key=f"store_phone2_{cid}")

    # 三大公版條款文字庫
    catering_default = "(1)使用本券請提前預約，預約時請告知使用本券及內容。\n(2)點餐前，請事先告知服務人員欲使用本券，結帳時出示本券畫面予櫃台掃碼使用。\n(3)飲料可加價升級特大杯或添加客製化等收費項目。\n(4)兌換數量依各門市現貨為準，若該門市已無存貨，請至鄰近門市兌換，或指定飲料可選擇更換其他系列品項。\n(5)本券平假日皆適用。\n(6)本券僅限於台灣區域使用且不適用於外送服務與網路訂餐。\n(7)本券商品不得與其他優惠合併使用且一桌限用一張。\n(8)本券商品無法累積會員積分。\n(9)本商品恕無法折抵商場停車。"
    beauty_default = "(1)使用本券請提前預約，預約時請告知使用本券及內容。\n(2)本券平假日皆適用。\n(3)本券商品無法累積會員積分。\n(4)本券不得指定美容師、療程房型；如您是敏感膚質，或有生理期、懷孕、心血管疾病等特殊身體狀況，請於課程開始前，主動告知美容師。\n(5)兌換數量依各門市現貨為準，若該門市已無存貨，請至鄰近門市兌換，或指定飲料可選擇更換其他系列品項。"
    entertainment_default = "(1)本優惠為單人使用，不建議 12 歲以下孩童參與。\n(2)本遊戲 4 人成團，未滿 8 人需與其他玩家併團遊戲。\n(3)本優惠活動時間約 100 分鐘(含講解)。"

    st.write("---")
    st.subheader("📦 商品品項與規格細節填寫")
    products_data = []

    for i in range(1, 6):
        with st.expander(f"🔹 點我填寫第 {i} 項商品內容", expanded=(i == 1)):
            p_c1, p_c2 = st.columns(2)
            with p_c1:
                p_code = st.text_input(f"商品序號 ({i})", placeholder="例如：001", key=f"p_code_{i}_{cid}")
                p_name = st.text_input(f"商品名稱/規格 ({i})", placeholder="例如：美式咖啡(大杯/冷)", key=f"p_name_{i}_{cid}")
                p_orig_price = st.text_input(f"商品原價 ({i})", placeholder="例如：150", key=f"p_orig_price_{i}_{cid}")
                p_disc_price = st.text_input(f"商品優惠價 ({i})", placeholder="例如：120", key=f"p_disc_price_{i}_{cid}")
                p_stock = st.text_input(f"庫存數 ({i})", placeholder="例如：1000", key=f"p_stock_{i}_{cid}")
                p_desc = st.text_area(f"商品描述及文案說明 ({i})", placeholder="關於商品的詳細說明", key=f"p_desc_{i}_{cid}", height=100)

            with p_c2:
                st.write("**📅 商品券兌換時間設定**")
                validity_type = st.radio(f"兌換類型 ({i})", ["常態商品 (自訂天數，最大150天)", "季節商品 (自訂截止日)"], key=f"v_type_{i}_{cid}")
                t_now = datetime.now()
                display_name = p_name if p_name else f"（第{i}項商品名稱/規格）"

                # 💡 核心優化 1 & 4：若選「常態商品」，強制鎖定格式二，自動套用星號 *
                if validity_type == "常態商品 (自訂天數，最大150天)":
                    p_valid_days = st.number_input(f"常態商品兌換天數 ({i})", min_value=1, max_value=150, value=50, key=f"p_valid_days_{i}_{cid}")
                    start_date_calc = t_now + timedelta(days=7)
                    end_date_calc = start_date_calc + timedelta(days=p_valid_days)
                    start_d = start_date_calc.strftime("%Y/%m/%d")
                    end_d = end_date_calc.strftime("%Y/%m/%d")
                    st.info(f"常態商品已自動計算（製作日+7天為開始日，共 {p_valid_days} 天）：{start_d} 至 {end_d}")
                    
                    final_validity_text = f"*本券可兌換{display_name}，兌換期間為購買當日起至{p_valid_days}日止。"
                    st.write("💡 常態商品固定套用格式二文案：")
                    st.code(final_validity_text, language="text")

                # 💡 核心優化 2 & 3 & 4：若選「季節商品」，更換欄位名稱，僅限格式一與格式三
                else:
                    vd1, vd2 = st.columns(2)
                    with vd1:
                        start_date_val = st.date_input(f"兌換起始日期 ({i})", value=t_now, key=f"s_date_{i}_{cid}")
                    with vd2:
                        end_date_val = st.date_input(f"兌換截止日期 ({i})", value=t_now + timedelta(days=45), key=f"e_date_{i}_{cid}")
                    
                    start_d = start_date_val.strftime("%Y/%m/%d")
                    end_d = end_date_val.strftime("%Y/%m/%d")
                    days_diff = (end_date_val - start_date_val).days
                    
                    if days_diff > 150:
                        st.error(f"⚠️ 警告：目前相差 {days_diff} 天已超過 150 天，注意！！兌換日不可超過150天。")
                    elif days_diff < 50:
                        st.warning(f"⚠️ 提示：目前相差 {days_diff} 天，低於常規 50 天（系統仍可正常導出）。")
                    else:
                        st.success(f"✅ 目前相差 {days_diff} 天，符合系統兌換 50~150天，注意！！兌換日不可超過150天。")
                    
                    format_choice = st.selectbox(
                        f"選擇要帶入的官方兌換時間文案格式 ({i})",
                        [
                            "格式一：*本券兌換期間為YYYY/MM/DD至YYYY/MM/DD止。",
                            "格式三：*本券可兌換商品名稱/規格，兌換期間為YYYY/MM/DD至YYYY/MM/DD止。"
                        ],
                        key=f"format_choice_{i}_{cid}"
                    )
                    
                    if "格式一" in format_choice:
                        final_validity_text = f"*本券兌換期間為{start_d}至{end_d}止。"
                    else:
                        final_validity_text = f"*本券可兌換{display_name}，兌換期間為{start_d}至{end_d}止。"
                    st.code(final_validity_text, language="text")

                st.write("**🛒 商品販售時間**")
                sell_type = st.radio(f"販售屬性 ({i})", ["常態商品 (填無)", "季節性商品 (填日期區間)"], key=f"sell_type_{i}_{cid}")
                
                if sell_type == "常態商品 (填無)":
                    final_sell_time = "無"
                else:
                    s1, s2 = st.columns(2)
                    with s1:
                        s_s = st.date_input(f"販售開始日 ({i})", value=t_now, key=f"s_s_{i}_{cid}")
                    with s2:
                        s_e = st.date_input(f"販售結束日 ({i})", value=t_now + timedelta(days=14), key=f"s_e_{i}_{cid}")
                    final_sell_time = f"{s_s.strftime('%Y/%m/%d')} 至 {s_e.strftime('%Y/%m/%d')}"

                st.write("**🏷️ 商品折扣時間**")
                discount_type = st.radio(f"折扣屬性 ({i})", ["原價販售 (填無)", "限定折扣 (填日期區間)"], key=f"discount_type_{i}_{cid}")
                
                if discount_type == "原價販售 (填無)":
                    final_discount_time = "無"
                else:
                    d1, d2 = st.columns(2)
                    with d1:
                        d_s = st.date_input(f"折扣開始日 ({i})", value=t_now, key=f"d_s_{i}_{cid}")
                    with d2:
                        d_e = st.date_input(f"折扣結束日 ({i})", value=t_now + timedelta(days=7), key=f"d_e_{i}_{cid}")
                    final_discount_time = f"{d_s.strftime('%Y/%m/%d')} 至 {d_e.strftime('%Y/%m/%d')}"

            st.write("---")
            st.write("**📜 使用條款說明公版選單**")
            terms_template = st.selectbox(f"選擇適用之產業條款範本 ({i})", ["餐飲公版條款", "美容公版條款", "休閒娛樂公版條款", "完全自訂空白"], key=f"terms_template_{i}_{cid}")
            terms_value = catering_default if terms_template == "餐飲公版條款" else (beauty_default if terms_template == "美容公版條款" else (entertainment_default if terms_template == "休閒娛樂公版條款" else ""))
            p_terms = st.text_area(f"編輯/修正使用條款內容 ({i})", value=terms_value, key=f"p_terms_{i}_{terms_template}_{cid}", height=120)

            if p_name:
                products_data.append({"code": p_code, "name": p_name, "orig_price": p_orig_price, "disc_price": p_disc_price, "stock": p_stock, "validity": final_validity_text, "desc": p_desc, "terms": p_terms, "sell_time": final_sell_time, "discount_time": final_discount_time})

    # 生成按鈕
    if st.button("🚀 生成完整票券 Excel 資料表", type="primary", key="btn_gen_excel"):
        if len(products_data) == 0:
            st.error("❌ 請至少填寫一項商品的『商品名稱/規格』才能進行 Excel 匯出！")
        else:
            b_display = brand_name if brand_name else "（請填寫 brand_name）"
            notices = f"1. 使用本券請至 {b_display} 直接出示本券掃碼兌換（請將螢幕亮度調到最大）。\n2. 本券恕不得更換現金及轉售。\n3. 使用本券時須符合本券載明之品項與規格。因購買時LINE Pay已開立發票給購買者，兌換時不另開立發票。商品兌換後，恕無法提供退貨及換貨。\n4. 商店僅提供兌換本券商品的服務，若對兌換之商品有任何問題請洽門市人員，其他本服務相關問題請聯繫連加網路商業股份有限公司（下稱LINE Pay）客服。\n5. 本券不記名，僅限兌換一次，不得重複使用，任何人持有皆可兌換，請自行妥善保管。\n6. 本券之兌換與銷售，恕不與商店所有折扣、優惠、各行銷活動合併使用。\n7. 有關本券之使用、兌換、取消及補發之條款及條件，及本服務之完整內容，請詳見「服務條款」。\n8. 本券如未於期限內兌換，費用將全額退款給原購買者。"
            issuer_info = "連加網路商業股份有限公司\n地址：臺北市南港區經貿二路121號18樓\n電話：02-3518-7600\n統編：24941093"
            guarantee_info = "本服務所發行之票券金額，皆自發行日起存入發行人於國泰世華商業銀行開立之信託帳戶，專款專用。所謂專用，係指供發行人履行交付商品或提供服務義務使用，前述信託期間自出售日起算至少一年。"
            provider_info = f"商店名稱：{b_display}\n地址：{store_address}\n電話：{store_phone1} / {store_phone2}\n統編：{store_tax_id}"

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "LINEPay新規文字上架表"
            ws.views.sheetView[0].showGridLines = True
            
            headers = ["商品序號", "品牌名稱", "商品名稱/規格", "商品券兌換時間說明", "商品描述及文案說明", "使用條款說明", "商品販售時間", "原價", "優惠價", "商品折扣時間", "庫存數", "禮券發行者(固定)", "履約保證(固定)", "注意事項說明", "實際商品(服務)提供者"]
            ws.append(headers)
            
            title_font, title_fill, title_align = Font(name="微軟正黑體", size=11, bold=True, color="FFFFFF"), PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid"), Alignment(horizontal="center", vertical="center", wrap_text=True)
            data_font, text_align, center_align = Font(name="微軟正黑體", size=10), Alignment(horizontal="left", vertical="top", wrap_text=True), Alignment(horizontal="center", vertical="top", wrap_text=True)
            thin_border = Border(left=Side(style="thin", color="D9D9D9"), right=Side(style="thin", color="D9D9D9"), top=Side(style="thin", color="D9D9D9"), bottom=Side(style="thin", color="D9D9D9"))
            
            col_widths = {'A': 15, 'B': 18, 'C': 30, 'D': 45, 'E': 45, 'F': 55, 'G': 30, 'H': 12, 'I': 12, 'J': 30, 'K': 12, 'L': 35, 'M': 45, 'N': 55, 'O': 40}
            for col_let, width in col_widths.items(): ws.column_dimensions[col_let].width = width
            
            ws.row_dimensions[1].height = 35
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font, cell.fill, cell.alignment, cell.border = title_font, title_fill, title_align, thin_border

            current_row = 2
            for p in products_data:
                ws.row_dimensions[current_row].height = 80
                ws.cell(row=current_row, column=1, value=p["code"])
                ws.cell(row=current_row, column=2, value=brand_name)
                ws.cell(row=current_row, column=3, value=p["name"])
                ws.cell(row=current_row, column=4, value=p["validity"])
                ws.cell(row=current_row, column=5, value=p["desc"])
                ws.cell(row=current_row, column=6, value=p["terms"])
                ws.cell(row=current_row, column=7, value=p["sell_time"])
                ws.cell(row=current_row, column=8, value=p["orig_price"])
                ws.cell(row=current_row, column=9, value=p["disc_price"])
                ws.cell(row=current_row, column=10, value=p["discount_time"])
                ws.cell(row=current_row, column=11, value=p["stock"])
                ws.cell(row=current_row, column=12, value=issuer_info)
                ws.cell(row=current_row, column=13, value=guarantee_info)
                ws.cell(row=current_row, column=14, value=notices)
                ws.cell(row=current_row, column=15, value=provider_info)

                for col_idx in range(1, len(headers) + 1):
                    cell = ws.cell(row=current_row, column=col_idx)
                    cell.font, cell.border = data_font, thin_border
                    cell.alignment = center_align if col_idx in [1, 2, 4, 7, 8, 9, 10, 11] else text_align
                current_row += 1

            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            st.session_state["final_excel_bytes"] = excel_buffer.getvalue()
            st.session_state["final_file_name"] = f"{brand_name if brand_name else 'LINEPay'}_新規純文字規格表.xlsx"

    if "final_excel_bytes" in st.session_state:
        st.success("🎉 純文字 Excel 生成成功！")
        st.download_button(label="💾 點我下載 2026 新規純文字 Excel 檔案", data=st.session_state["final_excel_bytes"], file_name=st.session_state["final_file_name"], mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)

# ==========================================
# 🖼️ 分頁二：全新官方圖片規格一鍵裁剪下載工具
# ==========================================
with tab2:
    st.title("🖼️ LINE Pay 官方圖片規格一鍵裁切下載工具")
    st.write("在這裡上傳的任何圖片，系統都會自動強制調整為 LINE Pay 官方要求的指定像素與比例。")
    
    st.write("---")
    
    # LOGO 部分
    st.subheader("① 店家形象 LOGO（自動調整為 300x300 px）")
    logo_tool_file = st.file_uploader("選擇 LOGO 圖片", type=["jpg", "jpeg", "png"], key=f"tool_logo_{cid}")
    if logo_tool_file:
        img_l = Image.open(logo_tool_file).resize((300, 300))
        st.image(img_l, width=150, caption="LOGO 預覽 (300x300 px)")
        l_buf = io.BytesIO()
        img_l.save(l_buf, format="PNG")
        st.download_button("💾 下載已裁切 LOGO", data=l_buf.getvalue(), file_name="LINEPay_LOGO_300x300.png", mime="image/png")

    st.write("---")

    # Banner 部分
    st.subheader("② 店家首頁 Banner 橫幅（自動調整為 750x454 px）")
    banner_tool_file = st.file_uploader("選擇 Banner 圖片", type=["jpg", "jpeg", "png"], key=f"tool_banner_{cid}")
    if banner_tool_file:
        img_b = Image.open(banner_tool_file).resize((750, 454))
        st.image(img_b, width=375, caption="Banner 預覽 (750x454 px)")
        b_buf = io.BytesIO()
        img_b.save(b_buf, format="PNG")
        st.download_button("💾 下載已裁切 Banner", data=b_buf.getvalue(), file_name="LINEPay_Banner_750x454.png", mime="image/png")

    st.write("---")

    # 商品圖部分 (5 ~ 10 張)
    st.subheader("③ 商品主要及附加照片（自動調整為 640x640 px，支援最多10張）")
    prod_tool_files = st.file_uploader("選擇 5 ~ 10 張商品照片（可複選多張上傳）", type=["jpg", "jpeg", "png"], accept_multiple_files=True, key=f"tool_prod_{cid}")
    
    if prod_tool_files:
        valid_prods = prod_tool_files[:10]  # 強制上限10張
        if len(valid_prods) < 5:
            st.warning(f"💡 目前上傳了 {len(valid_prods)} 張商品圖，官方建議上架準備 5 ~ 10 張照片。")
        else:
            st.success(f"✅ 已成功載入 {len(valid_prods)} 張商品照片！")
            
        zip_buffer = io.BytesIO()
        
        with zipfile.ZipFile(zip_buffer, "w") as zip_file:
            st.write("👇 **單張照片預覽與個別下載：**")
            
            for idx, p_file in enumerate(valid_prods):
                img_p = Image.open(p_file).resize((640, 640))
                p_buf = io.BytesIO()
                img_p.save(p_buf, format="PNG")
                p_bytes = p_buf.getvalue()
                
                zip_file.writestr(f"Product_Image_{idx+1}_640x640.png", p_bytes)
                
                p_col1, p_col2 = st.columns([1, 4])
                with p_col1:
                    st.image(img_p, width=120)
                with p_col2:
                    st.write(f"**商品照片 {idx+1}** (規格：640x640 px)")
                    st.download_button(f"📥 下載單張商品圖 {idx+1}", data=p_bytes, file_name=f"Product_Image_{idx+1}.png", mime="image/png", key=f"dl_single_p_{idx}_{cid}")
        
        st.write("---")
        st.write("🎁 **高效率大絕招：**")
        st.download_button(
            label="🔥 點我一鍵打包下載所有商品圖 (ZIP 壓縮檔)",
            data=zip_buffer.getvalue(),
            file_name="LINEPay_640x640_商品照片包.zip",
            mime="application/zip",
            type="primary",
            use_container_width=True
        )

# ==========================================
# 🧹 全局通用清除按鈕（放在網頁最底部）
# ==========================================
st.write("---")
if st.button("🧹 清除所有內容（做下一家店 / 清空照片）", use_container_width=True, key="btn_global_clear"):
    if "final_excel_bytes" in st.session_state: del st.session_state["final_excel_bytes"]
    if "final_file_name" in st.session_state: del st.session_state["final_file_name"]
    st.session_state["clear_id"] += 1
    st.rerun()
